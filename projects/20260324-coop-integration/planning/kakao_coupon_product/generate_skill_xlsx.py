"""
헬로우봇 카테고리별 스킬 목록 CSV → topic/intent별 멀티시트 xlsx 생성 스크립트

사용법:
    python3 generate_skill_xlsx.py

입력: 헬로우봇 카테고리별 스킬 목록 - 20260408 - 전체-미노출 제외 + 가격정보.csv
출력: 헬로우봇 카테고리별 스킬 목록 - 20260408 - 전체-미노출 제외 + 가격정보.xlsx

시트 구성:
  - 전체 (앱/웹 모두 노출인 항목만)
  - topic별 시트 (연애, 결혼, 일반운세, 총운, 재물금전, 학업직업, 자기탐구, 가족자녀, 기타)
  - 연애-intent 세부 시트 (intents를 | 기준 분리, 해당 키워드 포함 행 중복 포함)
"""

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

SCRIPT_DIR = Path(__file__).parent
INPUT_FILE = SCRIPT_DIR / "헬로우봇 카테고리별 스킬 목록 - 20260408 - 전체-미노출 제외 + 가격정보.csv"
OUTPUT_FILE = INPUT_FILE.with_suffix(".xlsx")


def clean(val):
    if isinstance(val, str):
        return ILLEGAL_CHARS_RE.sub("", val)
    return val


def add_sheet(wb, name, headers, data):
    safe = name[:31].replace("/", "_").replace("\\", "_")
    ws = wb.create_sheet(title=safe)
    ws.append(headers)
    for row in data:
        ws.append([clean(row.get(h, "")) for h in headers])


def main():
    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        all_rows = list(reader)

    # 앱/웹 노출여부 둘 다 '노출'인 항목만
    rows = [
        r
        for r in all_rows
        if r.get("앱 노출여부", "").strip() == "노출"
        and r.get("웹 노출여부", "").strip() == "노출"
    ]
    print(f"전체 {len(all_rows)}건 → 필터 후 {len(rows)}건 (제외 {len(all_rows) - len(rows)}건)")

    # topic별 그룹
    by_topic = defaultdict(list)
    for row in rows:
        by_topic[row.get("topic", "").strip() or "(없음)"].append(row)

    # 연애 → intent 키워드별 그룹 (| 분리, 중복 포함)
    love_by_intent = defaultdict(list)
    intent_set = set()
    for row in by_topic.get("연애", []):
        raw = row.get("intents", "").strip()
        if not raw or raw == "-":
            love_by_intent["-"].append(row)
            intent_set.add("-")
        else:
            for part in raw.split("|"):
                part = part.strip()
                love_by_intent[part].append(row)
                intent_set.add(part)

    wb = Workbook()
    wb.remove(wb.active)

    # 전체 시트
    add_sheet(wb, "전체", headers, rows)

    # topic별 시트
    topic_order = ["연애", "결혼", "일반운세", "총운", "재물금전", "학업직업", "자기탐구", "가족자녀", "기타", "-"]
    for topic in topic_order:
        if topic in by_topic:
            add_sheet(wb, topic if topic != "-" else "기타(-)", headers, by_topic[topic])

    # 연애-intent 세부 시트 (건수 내림차순)
    for intent in sorted(intent_set, key=lambda x: -len(love_by_intent[x])):
        add_sheet(wb, f"연애-{intent}", headers, love_by_intent[intent])

    wb.save(OUTPUT_FILE)

    print(f"\n생성 완료: {OUTPUT_FILE}")
    print(f"시트 수: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name}: {ws.max_row - 1}건")


if __name__ == "__main__":
    main()
